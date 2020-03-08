import sqlite3
import logging
from openpyxl import load_workbook

logging.disable(logging.DEBUG)
logging.basicConfig(
    level=logging.DEBUG, format=" %(asctime)s | %(levelname)s | %(message)s"
)
logging.basicConfig(
    level=logging.INFO, format=" %(asctime)s | %(levelname)s | %(message)s"
)


def update():
    """
    读取从MIS系统导出的Excel表中的内容
    清空数据库中今年的数据
    将今年的新数据写入数据库中
    """
    conn = sqlite3.connect(r'Data\2020年.db')
    logging.debug('数据库连接成功')
    cur = conn.cursor()

    table = '2020年'

    # 清空原数据库数据
    str_sql = f"DELETE FROM [{table}]"
    cur.execute(str_sql)
    conn.commit()
    logging.info("数据库数据清空完毕")

    # 读入Excel表格数据
    wb = load_workbook('云南分公司业务统计表.xlsx')
    ws = wb['page']

    logging.info("Excel 文件读入成功")
    logging.info(f"需要导入{ws.max_row}条数据")

    nrow = 1
    # 将Excel数据写入数据库中
    for row in ws.iter_rows(min_row=2,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=ws.max_column):
        str_sql = f"INSERT INTO '{table}' VALUES ("
        for v in row:
            str_sql += f"'{v.value}', "

        str_sql = str_sql[:-2] + ")"

        cur.execute(str_sql)

        nrow += 1
        if nrow % 500 == 0:
            logging.info(f'已导入 {nrow} / {ws.max_row} 条数据')

    logging.info('数据写入数据库完成')

    str_sql = f"DELETE FROM '{table}' \
                WHERE [车险/财产险/人身险] <> '车险' \
                AND [车险/财产险/人身险] <> '财产险' \
                AND [车险/财产险/人身险] <> '人身险'"
    cur.execute(str_sql)

    logging.info('非统计数据删除完成')

    conn.commit()

    logging.info("数据库事务提交完成")
    logging.info("数据库更新操作完成")

    cur.close()
    conn.close()
    print("-" * 60)


if __name__ == '__main__':
    update()
