# _*_ coding: utf-8 _*_

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.utils import get_column_letter, column_index_from_string
import logging

from datetime import date
from datetime import timedelta

from cell_style import cell_style
from my_date import MyDate
from stats import Stats

def kun_ming_week(sh, sql, risk, row_var = 0):
    """
    该函数负责将昆明周报数据写入Excel文件中，
    其中标题和表头在此函数中定义，而表中数据则通过SQL从数据库中提取
    sh：xlwt.Workbook.sheet变量，是待写入数据的excel工作表
    sql：mysql变量，负责执行SQL语句
    risk：str变量，记录险种信息
    ivar：int变量，记录数据写入时的行调整间隔
    """

    logging.debug(f'开始写入"{risk}"信息……')

    # 行变量，每一个统计表占14行，以此调整不同统计表的起始行
    row_var = (row_var-1) * 14

    # 设置车险汇总表标题
    sh.merge_cells(start_row=1 + row_var, start_column=1, end_row=1 + row_var, end_column=8)
    sh.cell(1+ row_var, 1).value = f"昆明地区机构{risk}保费汇总表"
    sh.cell(1+row_var, 1).font = Font(name="微软雅黑", size=14)

    logging.debug(f"{risk}保费汇总表表头写入完成")

    # 设置数据截止时间

    idate = MyDate()
    last_date = date.today() - timedelta(days = 1)
    date_str = "数据截止至 " + last_date.strftime("%Y-%m-%d")
    sh.merge_cells(start_row=2 + row_var, start_column=1, end_row=2 + row_var, end_column=8)
    sh.cell(2 + row_var, 1).value = date_str
    sh.cell(2 + row_var, 1).font = Font(name="微软雅黑", size=10)
    logging.debug(f"{risk}保费汇总表时间写入完成")
    
    # 设置表头
    header = ("机构名称", "周保费", "周环比", "周同比", "月保费", "月环比", "月同比", "年累计保费", "年同比")
    
    # enumerate函数可将元组或列表等变量转换为一个索引与值对应的元组列表，并同时返回索引和值两个数据
    for k, v in enumerate(header):
        column_letter = get_column_letter(k+1)
        if k == 0:
            # 第一列需要写入机构名称，宽度为14个字符
            sh.column_dimensions[column_letter].width = 14
        else:
            sh.column_dimensions[column_letter].width = 12

        sh.cell(3 + row_var, k + 1).value = v
        sh.cell(3 + row_var, k + 1).font = Font(name="微软雅黑", size=12, bold=True)
    
    ## 设置表数据
    #table_info = ("春怡雅苑", "香榭丽园", "百大国际", "宜良", "东川", "安宁", "春之城", "分公司本部", "飞航", "合计")

    ## 单一机构统计信息列表变量
    #ji_gou_data = []

    ## 数据表正文样式
    #text_style = cell_style(borders = True)
    #num_style = cell_style(borders = True, num_format = '0.00')
    #percent_style = cell_style(borders = True, num_format = '0.00%')

    ## 行计数变量
    #irow = 3
  
    #for ji_gou in table_info:

    #    logging.debug('写入"{0}"信息'.format(ji_gou))
    #    if ji_gou == '合计':
    #        data = Stats("昆明", risk, sql, "中心支公司")
    #        ji_gou_data = ("合计", data.last_week, data.last_week_huan_bi,  data.last_week_tong_bi, data.this_month, data.month_huan_bi, 
    #                    data.month_tong_bi, data.this_year, data.year_tong_bi)
    #    else:
    #        data = Stats(ji_gou, risk, sql, "机构")
    #        ji_gou_data = (data.jian_cheng, data.last_week, data.last_week_huan_bi,  data.last_week_tong_bi, data.this_month, data.month_huan_bi, 
    #                        data.month_tong_bi, data.this_year, data.year_tong_bi)
        
    #    for k, v in enumerate(ji_gou_data):
    #        if k == 0:
    #            sh.row(irow + row_var).write(k, v, text_style)
    #        elif k == 1 or k ==4 or k == 7:
    #            sh.row(irow + row_var).write(k, v, num_style)
    #        else:
    #            sh.row(irow + row_var).write(k, v, percent_style)
       
    #    irow += 1
    #logging.debug('"{0}"信息写入完成……\n'.format(risk))

