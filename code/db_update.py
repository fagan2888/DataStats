import sqlite3
import logging
from openpyxl import load_workbook


logging.disable(logging.NOTSET)
logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s | %(levelname)s | %(message)s')


def update(conn, cur):
    """
    读取从MIS系统导出的Excel表中的内容
    清空数据库中今年的数据
    将今年的新数据写入数据库中
    """

    # 载入excel中的业务清单表,并对后期 sql语句数据插入做准备
    wb = load_workbook('云南分公司业务统计表', read_only=True)
    ws = wb['page']

    logging.debug("Excel 文件读入成功")

    # 清空原数据库数据
    str_sql = "DELETE FROM [2015年]"
    cur.execute(str_sql)
    conn.commit()

    logging.debug("数据库数据清空完毕")

    for row in ws.iter_rows(min_row=2,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=ws.max_column,
                            values_only=True):
        str_sql = "INSERT INTO [2015年] VALUES ("
        for value in row:
            str_sql += f" ,'{value}'"

        str_sql += ")"

    conn.commit()

    logging.debug("数据库数据插入完成")


if __name__ == '__main__':
    conn = sqlite3.connect()
    cur = conn.cursor()

    update(conn, cur)

    cur.close()
    conn.colse()
