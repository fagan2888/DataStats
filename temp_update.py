import sqlite3
import logging
from openpyxl import load_workbook


def update(tb_name=None, back=True):
    """
    读取从MIS系统导出的Excel表中的内容
    清空数据库中今年的数据
    将今年的新数据写入数据库中
    """
    conn = sqlite3.connect(r'Data\data.db')
    logging.debug('数据库连接成功')
    cur = conn.cursor()

    if tb_name is None:
        table = '2020年开门红任务'
    else:
        table = tb_name

    if back is True:
        path = f"Back\\{table}.xlsx"
    else:
        path = f"{table}.xlsx"

    # 清空原数据库数据
    str_sql = f"DELETE FROM [{table}]"
    cur.execute(str_sql)
    conn.commit()
    logging.debug("数据库数据清空完毕")

    # 读入Excel表格数据
    wb = load_workbook(path)
    ws = wb.active

    logging.debug("Excel 文件读入成功")
    logging.debug(f"需要导入{ws.max_row}条数据")

    nrow = 1
    # 将Excel数据写入数据库中
    for row in ws.iter_rows(min_row=2,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=ws.max_column):
        str_sql = f"INSERT INTO '{table}' VALUES ("
        for v in row:
            str_sql += f"'{v.value}', "

        str_sql = str_sql[:-2] + ")"

        cur.execute(str_sql)

        nrow += 1
        if nrow % 1000 == 0:
            logging.debug(f'已导入 {nrow} / {ws.max_row} 条数据')

    logging.debug('数据写入数据库完成')

    conn.commit()

    logging.debug("数据库事务提交完成")

    cur.close()
    conn.close()


if __name__ == '__main__':
    update()
