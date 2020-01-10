import os
import sys
import shutil
import logging
from openpyxl import Workbook
from datetime import datetime

from zhong_zhi_write import zhong_zhi_write
from ji_gou_write import ji_gou_write
from tuan_dui_write import tuan_dui_write
from style import style
from update import update


logging.disable(logging.NOTSET)
logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s | %(levelname)s | %(message)s')

# # 调用数据库更新函数
# update()

wb = Workbook()

# 调用样式函数初始化数字样式
style(wb)

# # 写入三级机构中心支公司数据表
# ws = wb.active
# ws.title = '三级机构数据统计表'
# zhong_zhi_write(ws)

# # 写入四级机构数据表
# ws = wb.create_sheet(title='四级机构数据统计表')
# ji_gou_write(ws)

# 写入内部团队数据表
ws = wb.create_sheet(title='内部团队数据统计表')
tuan_dui_write(ws)

wb.save('2020年机构数据统计表.xlsx')
