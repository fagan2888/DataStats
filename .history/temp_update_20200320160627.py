import sqlite3
import logging
from openpyxl import load_workbook

logging.disable(logging.NOTSET)
logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s | %(levelname)s | %(message)s')


def update(tb_name=None, back=True):
    """
    读取从MIS系统导出的Excel表中的内容
    清空数据库中今年的数据
    将今年的新数据写入数据库中
    """
    table = tb_name

    conn = sqlite3.connect(f'Data\\{table}.db')
    logging.debug('数据库连接成功')
    cur = conn.cursor()

    if back is True:
        path = f"Back\\{table}.xlsx"
    else:
        path = f"{table}.xlsx"

    # # 清空原数据库数据
    # str_sql = f"DELETE FROM [{table}]"
    # cur.execute(str_sql)
    # conn.commit()
    # logging.debug("数据库数据清空完毕")

    # 读入Excel表格数据
    wb = load_workbook(path, read_on)
    ws = wb.active

    logging.debug("Excel 文件读入成功")
    logging.debug(f"需要导入{ws.max_row}条数据")

    nrow = 1
    # 将Excel数据写入数据库中
    for row in ws.iter_rows(min_row=2,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=ws.max_column):
        str_sql = f"INSERT OR IGNORE INTO '{table}' VALUES ("
        for v in row:
            str_sql += f"'{v.value}', "

        str_sql = str_sql[:-2] + ")"

        cur.execute(str_sql)

        nrow += 1
        if nrow % 5000 == 0:
            logging.debug(f'已导入 {nrow} / {ws.max_row} 条数据')

    logging.debug('数据写入数据库完成')

    conn.commit()

    logging.debug("数据库事务提交完成")

    cur.close()
    conn.close()


if __name__ == '__main__':
    ta_name = "车险机动车类型"
    back = False
    update(tb_name=ta_name, back=back)
    ta_name = "掌上宝APP出单统计"
    back = False
    update(tb_name=ta_name, back=back)
    ta_name = "车险清单"
    back = False
    update(tb_name=ta_name, back=back)
